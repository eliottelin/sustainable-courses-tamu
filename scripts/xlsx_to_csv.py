#!/usr/bin/env python3
"""
OSCE Syllabi Review XLSX → Site CSV Converter
---------------------------------------------
Converts a syllabi review XLSX (standard STARS format) into one CSV per
semester, ready to drop into the data/ folder of the sustainable course guide.

Usage:
  python3 scripts/xlsx_to_csv.py ReviewFile.xlsx
  python3 scripts/xlsx_to_csv.py data/*.xlsx
  python3 scripts/xlsx_to_csv.py ReviewFile.xlsx --semester "Fall 2026" --output "data/Fall 2026.csv"

The script:
  - Reads all semester/level sheets (skips Summary)
  - Derives the semester and level from each sheet's name when it follows the
    "<Semester> - <Level>" convention (e.g. "Fall 2026 - Undergraduate"). If a
    sheet's name doesn't carry a semester (e.g. a flat "Undergraduate" sheet),
    the --semester argument is used instead.
  - If --semester is given, only sheets for that semester are converted, and
    the result is written to --output (or "data/<semester>.csv").
  - If --semester is omitted, every semester found in the workbooks is
    converted, each to its own "data/<semester>.csv".
  - Expands cross-listed courses (AFST/ENGL → two rows). Handles both the old
    "Department(s)=short codes, Course=number" layout and the newer
    "Department(s)=full name(s), Course='CODE[/CODE...] number'" layout,
    surfacing full department name(s) in the 'department_full' column.
  - Converts 'yes' SDG cells to comma-separated SDG numbers
  - If the XLSX already has a 'Syllabus URL' column, it maps it automatically
  - When multiple xlsx files are passed, any sheet in the "raw" layout
    (headers include SUBJECT, COURSE, URL — e.g. a CRN/keyword-scan export)
    is used to build a syllabus URL lookup keyed by (subject, course number),
    which fills in syllabus_url for review rows that don't have their own URL.
  - Sheets that don't match either the review layout ('Course Title' header)
    or the raw lookup layout are silently skipped, so unrelated workbooks in
    data/ are safe.
"""
import sys, csv, re, argparse
import openpyxl

def parse_args():
    p = argparse.ArgumentParser(description='Convert OSCE review XLSX to site CSV')
    p.add_argument('xlsx', nargs='+', help='Path(s) to the syllabi review XLSX file(s), e.g. data/*.xlsx')
    p.add_argument('--semester', help='Only convert this semester, e.g. "Fall 2026" (default: convert every semester found)')
    p.add_argument('--output', help='Output CSV path (only valid with --semester and a single input file); default "data/<semester>.csv"')
    return p.parse_args()

def infer_level(text):
    return 'Undergraduate' if 'undergrad' in text.lower() else ('Graduate' if 'grad' in text.lower() else 'Undergraduate')

def parse_sheet_name(sheet_name, fallback_semester):
    """Derive (semester, level) from a sheet name like 'Fall 2026 - Undergraduate'."""
    if ' - ' in sheet_name:
        sem_part, level_part = sheet_name.split(' - ', 1)
        return sem_part.strip(), infer_level(level_part)
    return fallback_semester, infer_level(sheet_name)

def normalize_header(row, upper=False):
    """Safely stringify/strip a header row's cells (openpyxl returns None for blanks)."""
    return [(str(h).strip().upper() if upper else str(h).strip()) if h else '' for h in row]

def build_syllabus_lookup(workbooks):
    """Scan every sheet of every already-opened workbook for the raw
    CRN/keyword-scan layout (headers include SUBJECT, COURSE, URL) and build a
    (subject, course_number) -> url lookup used to fill in syllabus_url for
    review rows that don't carry their own URL column."""
    lookup = {}
    for wb in workbooks:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
            if not header_row:
                continue
            header = normalize_header(header_row, upper=True)
            if not ('SUBJECT' in header and 'COURSE' in header and 'URL' in header):
                continue
            subj_col, course_col, url_col = header.index('SUBJECT'), header.index('COURSE'), header.index('URL')
            for row in ws.iter_rows(min_row=2, values_only=True):
                if len(row) <= max(subj_col, course_col, url_col):
                    continue
                subject, course, url = row[subj_col], row[course_col], row[url_col]
                if not subject or not course or not url:
                    continue
                key = (str(subject).strip().upper(), str(course).strip())
                lookup.setdefault(key, str(url).strip())
    return lookup

def parse_course_cell(raw):
    """'Course' cell may be old-style ('393') or new-style ('AFST/ENGL 393').
    Returns (codes, number) - codes is [] for old-style (codes come from the
    Department(s) cell instead)."""
    raw = raw.strip()
    head, _, tail = raw.rpartition(' ')
    if head and any(ch.isalpha() for ch in head):
        return [c.strip() for c in head.split('/') if c.strip()], tail.strip()
    return [], raw

def parse_dept_names(raw):
    """Split a Department(s) cell into full names. New-style workbooks join
    cross-listed names with 'x'/'X' and irregular spacing; old-style is a
    single short code with no separator."""
    parts = re.split(r'\s+[xX]\s+', raw.strip())
    return [p.strip() for p in parts if p.strip()]

def parse_sheet(ws, level, semester, syllabus_lookup=None):
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    if not rows:
        return []
    # Find header row (first row with 'Course Title')
    header_idx = None
    for i, row in enumerate(rows):
        if row[0] and 'course title' in str(row[0]).lower():
            header_idx = i
            break
    if header_idx is None:
        return []
    header = normalize_header(rows[header_idx])

    # Detect optional syllabus URL column
    url_col = None
    for i, h in enumerate(header):
        if 'syllabus' in h.lower() and ('url' in h.lower() or 'link' in h.lower()):
            url_col = i
            break

    out = []
    for row in rows[header_idx+1:]:
        if not row[0]:
            continue
        title      = str(row[0]).strip()
        depts_raw  = str(row[1]).strip() if len(row) > 1 and row[1] else ''
        course_raw = str(row[2]).strip() if len(row) > 2 and row[2] else ''
        # level from col 3 (UG/G) or passed in
        description = str(row[4]).strip() if len(row) > 4 and row[4] else ''
        cls         = str(row[5]).strip() if len(row) > 5 and row[5] else ''

        # SDGs: columns 6–22 (Goals 1–17), 'yes' = active
        sdgs = []
        for i in range(17):
            idx = 6 + i
            val = row[idx] if idx < len(row) else None
            if val and str(val).strip().lower() == 'yes':
                sdgs.append(i + 1)

        syllabus_url = ''
        if url_col is not None and url_col < len(row) and row[url_col]:
            syllabus_url = str(row[url_col]).strip()

        codes_from_course, course_num = parse_course_cell(course_raw)
        if codes_from_course:
            # New-style row: dept codes live in the Course cell, full names in
            # the Department(s) cell (joined by 'x'/'X', not always 1:1 with codes).
            dept_list = codes_from_course
            dept_names = parse_dept_names(depts_raw)
            if len(dept_names) == len(dept_list):
                full_names = dept_names
            else:
                joined = ' / '.join(dept_names) if dept_names else ''
                full_names = [joined] * len(dept_list)
        else:
            # Old-style row: short code(s) come from the Department(s) cell itself.
            dept_list = [d.strip() for d in depts_raw.split('/') if d.strip()]
            full_names = [''] * len(dept_list)

        if not dept_list:
            dept_list = ['']
            full_names = ['']

        all_departments = '/'.join(dept_list)

        for dept, full_name in zip(dept_list, full_names):
            row_url = syllabus_url
            if not row_url and syllabus_lookup and course_num:
                row_url = syllabus_lookup.get((dept.strip().upper(), course_num), '')
            out.append({
                'course_title':    title,
                'department':      dept,
                'department_full': full_name,
                'all_departments': all_departments,
                'course_number':   course_num,
                'level':           level,
                'description':     description,
                'classification':  cls,
                'sdgs':            ','.join(str(s) for s in sdgs),
                'syllabus_url':    row_url,
                'semester':        semester,
            })
    return out

def main():
    args = parse_args()
    if args.output and not args.semester:
        sys.exit('--output requires --semester (ambiguous when converting multiple semesters)')
    if args.output and len(args.xlsx) > 1:
        sys.exit('--output is only valid with a single input file')

    workbooks = {path: openpyxl.load_workbook(path, read_only=True) for path in args.xlsx}
    syllabus_lookup = build_syllabus_lookup(workbooks.values())

    by_semester = {}
    sheet_sources = {}  # (semester, level) -> path of the first file that contributed rows for it
    for path, wb in workbooks.items():
        for sheet_name in wb.sheetnames:
            if sheet_name.lower() == 'summary':
                continue
            semester, level = parse_sheet_name(sheet_name, args.semester)
            if not semester:
                print(f'  Skipping sheet "{sheet_name}" in {path}: no semester in sheet name and --semester not provided')
                continue
            if args.semester and semester != args.semester:
                continue
            rows = parse_sheet(wb[sheet_name], level, semester, syllabus_lookup)
            if rows:
                key = (semester, level)
                if key in sheet_sources and sheet_sources[key] != path:
                    print(f'  WARNING: "{semester} - {level}" rows found in both {sheet_sources[key]} and {path} '
                          f'- combining, but this usually means an old backup workbook was left in data/ '
                          f'unarchived. Move outdated review workbooks to data/archive/.')
                else:
                    sheet_sources[key] = path
            by_semester.setdefault(semester, []).extend(rows)
            print(f'  {path} :: {sheet_name}: {len(rows)} rows -> semester "{semester}", level "{level}"')

    if not by_semester:
        print('\nNo matching rows found; nothing written.')
        return

    fieldnames = ['course_title','department','department_full','all_departments','course_number',
                  'level','description','classification','sdgs','syllabus_url','semester']
    for semester, rows in by_semester.items():
        if not rows:
            print(f'  Skipping "{semester}": 0 rows, not writing an empty CSV')
            continue
        out_path = args.output if args.output else f'data/{semester}.csv'
        with open(out_path, 'w', newline='', encoding='utf-8') as f:
            w = csv.DictWriter(f, fieldnames=fieldnames)
            w.writeheader()
            w.writerows(rows)
        print(f'Done: {len(rows)} rows written to {out_path}')

if __name__ == '__main__':
    main()
